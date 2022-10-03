import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import (
    PieChart,
    Reference
)
import pandas as pd


class ExcelBook():
    def __init__(self, filePath):
        self.filePath = filePath
        self.workBook = None
        self.workSheet = None
        self.workSheetName = None
        self.workSheetList = []
    
    def fileExists(self):
        
        if os.path.isfile(self.filePath):
            return True
        else:
            return False
    
    def getFile(self, workSheetName):

        if self.fileExists():
            
            self.workBook = load_workbook(filename = self.filePath)
            self.workSheetList = self.workBook.sheetnames

            self.workSheetName = workSheetName

            if workSheetName in self.workSheetList:
                
                self.workSheet = self.workBook[self.workSheetName]

            else:
                self.workSheet = self.workBook.create_sheet(self.workSheetName, 0)
           
            return True

        else:
            return False


    def createFile(self, workSheetName):
        if not self.fileExists():
            try:
                self.workBook = Workbook()
                self.workSheet = self.workBook.create_sheet(workSheetName, 0)

                return True
            except:
                print("ERROR: Unable to create Excel file")
                return False
        else:
            print("INFO: File exists: {}".format(self.fileExists()))
            return False

    def addData(self, data):
        for row in data:
            self.workSheet.append(row)
        self.workBook.save(filename=self.filePath)

    def addDataFromDataFrame(self, df):

        # NOTE: will append row by row if content is present
        for r in dataframe_to_rows(df, index=True, header=True):
            self.workSheet.append(r)
        
        for cell in self.workSheet['A'] + self.workSheet[1]:
            cell.style = 'Pandas'

        self.workBook.save(filename=self.filePath)

    def getDataRowByRow(self):

        data = []

        for row in self.workSheet.iter_rows(values_only=True):
            data.append(row)

        return data

    def getDataAsDataFrame(self):
        
        df = pd.DataFrame(self.workSheet.values)

        return df
    
    def createPieChart(self, data):
        pie = PieChart()
        labels = Reference(self.workSheet, min_col=1, min_row=2, max_row=5)
        data = Reference(self.workSheet, min_col=2, min_row=1, max_row=5)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Market Share of Tech by Category"

        self.workSheet.add_chart(pie, "D1")
        self.workBook.save(filename=self.filePath)


    
